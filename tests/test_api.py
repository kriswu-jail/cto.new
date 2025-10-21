from pathlib import Path

from fastapi.testclient import TestClient

from app.main import app, STORAGE_DIR


def test_api_export_png(tmp_path: Path):
    client = TestClient(app)

    # Create a small xlsx
    import openpyxl

    xlsx = tmp_path / "upload.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Hello"
    wb.create_sheet("Second")
    wb.save(str(xlsx))

    with open(xlsx, "rb") as f:
        resp = client.post(
            "/export/png",
            files={"file": (xlsx.name, f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"dpi": 72},
        )
    assert resp.status_code == 200
    data = resp.json()
    assert data["count"] == 2
    assert data["images"] and isinstance(data["images"], list)
    assert data["zip"], "Zip URL should be provided for multiple images"

    # Check that the assets exist on disk
    job_id = data["jobId"]
    job_dir = STORAGE_DIR / job_id
    assert job_dir.exists()
    for rel in data["images"]:
        p = job_dir / Path(rel).name
        assert p.exists(), f"Asset missing: {p}"
