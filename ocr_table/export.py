from __future__ import annotations
from typing import List, Optional

from .models import Table


def tables_to_dataframes(tables: List[Table]):
    try:
        import pandas as pd  # type: ignore
    except Exception as exc:
        raise ImportError("pandas is required to convert tables to DataFrame") from exc
    return [t.to_dataframe() for t in tables]


def export_tables_to_xlsx(
    tables: List[Table],
    output_path: str,
    sheet_names: Optional[List[str]] = None,
) -> None:
    try:
        import pandas as pd  # type: ignore
    except Exception as exc:  # pragma: no cover - optional
        raise ImportError("pandas is required for XLSX export") from exc

    engine = None
    try:
        import xlsxwriter  # type: ignore
        engine = "xlsxwriter"
    except Exception:
        try:
            import openpyxl  # type: ignore
            engine = "openpyxl"
        except Exception as exc:
            raise ImportError("xlsxwriter or openpyxl required for XLSX export") from exc

    writer = pd.ExcelWriter(output_path, engine=engine)  # type: ignore
    try:
        for idx, table in enumerate(tables):
            df = table.to_dataframe()
            sheet_name = (
                sheet_names[idx]
                if sheet_names and idx < len(sheet_names)
                else f"Table{idx+1}"
            )
            df.to_excel(writer, index=False, header=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            # Apply merges and simple border formatting if supported
            try:
                if engine == "xlsxwriter":
                    fmt = writer.book.add_format({"border": 1})
                    # Apply border format to all cells written
                    for r in range(table.rows):
                        for c in range(table.cols):
                            ws.write(r, c, table.grid[r][c], fmt)
                    for m in table.merges:
                        ws.merge_range(m.first_row, m.first_col, m.last_row, m.last_col, table.grid[m.first_row][m.first_col], fmt)
                elif engine == "openpyxl":
                    from openpyxl.styles import Border, Side
                    from openpyxl.utils import get_column_letter

                    thin = Side(border_style="thin", color="000000")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    ws_obj = ws  # openpyxl worksheet
                    # Apply values and borders
                    for r in range(1, table.rows + 1):
                        for c in range(1, table.cols + 1):
                            cell = ws_obj.cell(row=r, column=c)
                            cell.value = table.grid[r - 1][c - 1]
                            cell.border = border
                    # Merges (openpyxl uses 1-based inclusive coordinates)
                    for m in table.merges:
                        ws_obj.merge_cells(
                            start_row=m.first_row + 1,
                            start_column=m.first_col + 1,
                            end_row=m.last_row + 1,
                            end_column=m.last_col + 1,
                        )
            except Exception:
                # Best-effort formatting; ignore if engine doesn't support operations
                pass
    finally:
        writer.close()
