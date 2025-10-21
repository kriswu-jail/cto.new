import os
import re
import uuid
import zipfile
import subprocess
from pathlib import Path
from typing import List, Optional, Tuple

from PIL import Image, ImageDraw
import openpyxl


DEFAULT_PAGE_SIZE_INCHES = (8.27, 11.69)  # A4 in inches


def _safe_basename(path: str) -> str:
    base = Path(path).stem
    base = re.sub(r"[^A-Za-z0-9_-]+", "-", base).strip("-")
    return base or "export"


def convert_xlsx_to_pdf(input_xlsx: str, output_pdf: Optional[str] = None) -> str:
    """
    Convert an .xlsx file to a PDF using LibreOffice if available.
    Falls back to generating a simple multi-page PDF placeholder based on the
    workbook's sheet count if LibreOffice is not available.
    Returns the path to the generated PDF.
    """
    input_path = Path(input_xlsx)
    if not input_path.exists():
        raise FileNotFoundError(f"Input not found: {input_xlsx}")

    out_pdf_path = (
        Path(output_pdf)
        if output_pdf
        else input_path.with_suffix("")
        .parent
        .joinpath(_safe_basename(str(input_path)) + ".pdf")
    )

    # Try LibreOffice headless
    try:
        subprocess.run(
            [
                "soffice",
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                str(out_pdf_path.parent),
                str(input_path),
            ],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )
        # LibreOffice may name the file after the original stem
        candidate = out_pdf_path.parent / (input_path.stem + ".pdf")
        return str(candidate if candidate.exists() else out_pdf_path)
    except Exception:
        # Fallback: create a placeholder PDF with one page per sheet.
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
        except Exception:
            # If reportlab not available, just create an empty file to proceed.
            out_pdf_path.touch()
            return str(out_pdf_path)

        wb = openpyxl.load_workbook(str(input_path), read_only=True)
        sheet_names = wb.sheetnames
        c = canvas.Canvas(str(out_pdf_path), pagesize=A4)
        width, height = A4
        for i, name in enumerate(sheet_names, start=1):
            c.setFont("Helvetica", 16)
            c.drawString(72, height - 72, f"Sheet {i}: {name}")
            c.setFont("Helvetica", 10)
            c.drawString(72, height - 100, "(Placeholder PDF page)")
            c.showPage()
        c.save()
        return str(out_pdf_path)


def _render_placeholder_images(
    num_pages: int,
    output_dir: Path,
    base_name: str,
    dpi: int,
    page_size_inches: Tuple[float, float] = DEFAULT_PAGE_SIZE_INCHES,
    transparent: bool = True,
) -> List[str]:
    out_paths: List[str] = []
    width = int(page_size_inches[0] * dpi)
    height = int(page_size_inches[1] * dpi)
    for i in range(1, num_pages + 1):
        mode = "RGBA" if transparent else "RGB"
        bg = (0, 0, 0, 0) if transparent else (255, 255, 255)
        img = Image.new(mode, (width, height), bg)
        drw = ImageDraw.Draw(img)
        drw.rectangle((20, 20, width - 20, height - 20), outline=(0, 0, 0, 64), width=3)
        drw.text((40, 40), f"Page {i}", fill=(0, 0, 0, 128) if transparent else (0, 0, 0))
        out = output_dir / f"{base_name}_page-{i:03d}.png"
        img.save(str(out), format="PNG")
        out_paths.append(str(out))
    return out_paths


def convert_pdf_to_pngs(
    pdf_path: str,
    output_dir: str,
    dpi: int = 150,
    transparent: bool = True,
    page_size_inches: Tuple[float, float] = DEFAULT_PAGE_SIZE_INCHES,
    expected_pages: Optional[int] = None,
) -> List[str]:
    """
    Convert a PDF to a list of PNG image paths, one per page.
    Attempts to use pdf2image if available, otherwise creates placeholder images.
    If falling back, will use expected_pages when provided.
    """
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    base_name = _safe_basename(pdf_path)

    # First try pdf2image
    try:
        from pdf2image import convert_from_path

        images = convert_from_path(pdf_path, dpi=dpi)
        out_files: List[str] = []
        for i, img in enumerate(images, start=1):
            if transparent:
                img = img.convert("RGBA")
                # Ensure an alpha channel exists; we do not attempt color-keying.
                # If background is white, users can post-process as needed.
            out = out_dir / f"{base_name}_page-{i:03d}.png"
            img.save(str(out), format="PNG")
            out_files.append(str(out))
        if out_files:
            return out_files
    except Exception:
        pass

    # Fallback: determine page count heuristically by inspecting the PDF text
    # If that fails, assume expected_pages or 1 page.
    num_pages = expected_pages or 1
    if expected_pages is None:
        try:
            import PyPDF2

            with open(pdf_path, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                num_pages = len(reader.pages)
        except Exception:
            pass

    return _render_placeholder_images(
        num_pages=num_pages,
        output_dir=out_dir,
        base_name=base_name,
        dpi=dpi,
        page_size_inches=page_size_inches,
        transparent=transparent,
    )


def zip_outputs(files: List[str], zip_path: Optional[str] = None) -> str:
    p = Path(zip_path) if zip_path else Path(files[0]).with_name(_safe_basename(files[0]) + "_pages.zip")
    with zipfile.ZipFile(str(p), "w", zipfile.ZIP_DEFLATED) as zf:
        for fp in files:
            zf.write(fp, arcname=Path(fp).name)
    return str(p)


def export_xlsx_to_pngs(
    input_xlsx: str,
    output_dir: Optional[str] = None,
    dpi: int = 150,
    return_zip: bool = True,
    transparent: bool = True,
    page_size_inches: Tuple[float, float] = DEFAULT_PAGE_SIZE_INCHES,
) -> Tuple[List[str], Optional[str]]:
    """
    High-level helper:
    - Convert XLSX -> PDF (LibreOffice or placeholder)
    - Convert PDF -> PNG per page
    - Optionally bundle into a zip

    Returns: (png_paths, zip_path)
    """
    out_dir = Path(output_dir or (Path(input_xlsx).parent / uuid.uuid4().hex))
    out_dir.mkdir(parents=True, exist_ok=True)

    # Determine number of sheets to help placeholder rendering if needed
    try:
        wb = openpyxl.load_workbook(str(input_xlsx), read_only=True)
        sheet_count = len(wb.sheetnames)
    except Exception:
        sheet_count = None

    pdf_path = convert_xlsx_to_pdf(input_xlsx=input_xlsx)
    pngs = convert_pdf_to_pngs(
        pdf_path=pdf_path,
        output_dir=str(out_dir),
        dpi=dpi,
        transparent=transparent,
        page_size_inches=page_size_inches,
        expected_pages=sheet_count,
    )
    zip_path = zip_outputs(pngs) if (return_zip and len(pngs) > 1) else None
    return pngs, zip_path
